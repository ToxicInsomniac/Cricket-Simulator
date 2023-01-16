from random import choice, randint
import time

def slowprint(text,n=0.05):
    import sys, time
    for i in text:
        sys.stdout.write(i)
        time.sleep(n)
    print()
        
class Player:
    def __init__(self,name):
        self.name = name
        
        # Batting Attributes
        self.rs, self.bf, self._4, self._6, self.dp = 0, 0, 0, 0, 0
        self.F50, self.F100 = 'x', 'x'
        self.status = 'Y2B'
        self.contribution = [0,0]

        # Bowling Attributes
        self.bb, self.md, self.db, self.wt, self.rg, self.ex = 0, 0, 0, 0, 0, 0

    def sr(self):
        try:
            return str(round((self.rs*100)/self.bf,2))
        except:
            return 0.0
        
    def ob(self):
        try:
            return str(self.bb//6 + (self.bb%6)/10)
        except:
            return 0.0
        
    def eco(self):
        try:
            return str(round((self.rg*6)/self.bb,2))
        except:
            return 0.0

def check_player(message, XI, error):
    Team_Players = [player.name for player in XI]
    while True:
        p = input(message)
        if p in Team_Players:
            break
        else:
            print(error)
    return XI[Team_Players.index(p)]
    
            
def StrikeChange(S,NS):
    S.status = 'NS'
    NS.status = 'S'
    
    
def coin_toss(Team1, Team2, CT, XI1, XI2, Squad1, Squad2):
    coin = choice(['Heads','Tails'])
    print('It lands on', coin)
    CT.append(coin)
    
    if CT[0] == CT[-1]:
        print(Team1,'won the toss....')
        CT.append(Team1)
        ___ = 0
        
    else:
        print(Team2,'won the toss....')
        CT.append(Team2)
        ___ = 1

    while True:
        Decision = input('....and choose to (Bat/Bowl)-- ')
        insults = ["I don't even know why I try with you. It's Bat. OR. Bowl.","Repeat it with me. bat. OR! bowl.","I give up.",
                   "Of course it's case sensitive dumbass."]
        if Decision not in ["Bat","Bowl"]:
            slowprint(choice(insults))
            slowprint("Let's try that again, shall we?")
        else:
            break
    CT.append(Decision)
    XI = choose_to(CT[-1],___, XI1, XI2, Squad1, Squad2, Team1, Team2)
    return XI


            
def choose_to(C, n, XI1, XI2, Squad1, Squad2, Team1, Team2):
    if n == 0:
        if C == 'Bat':
            return [[XI1, XI2], [Squad1, Squad2]]
        else:
            return [[XI2, XI1], [Squad2, Squad1]]
    elif n==1:
        if C == 'Bat':
            return [[XI2, XI1], [Squad2, Squad1]]
        else:          
            return [[XI1, XI2], [Squad1, Squad2]]


def buff_nerf(value,net_buff_nerf):
    value += round((value*net_buff_nerf)/100)
    return value

def prob(match,balls):
    BD, RN, WTA, RCA = 0, 0, 0, 0
    
    if match == 'T20':
        freq = [2590,3610,810,120,1030,460,590,490,80,40,180]
    if match == 'ODI':
        freq = [4650,2520,960,170,710,190,340,230,40,80,100]
    if match == 'SuperOver':
        freq = [1040,4280,1730,120,1030,390,690,470,210,90,50]
    if match == 'Test':
        freq = [7160,1340,380,110,690,50,170,30,10,40,20]
       
    # freq=[f0,f1,f2,f3,f4,fOut,f6,fwd,fnb,fb,flb]
    buffs = {0: round(RCA/2 + WTA/2 - BD/4 - RN/4 , 2), 1: round(RN - RCA/4, 2), 2: round(RN - RCA/3, 2), 3: round(RN/3 - RCA/2 - WTA/2,2),
             4: round(BD - RCA/2, 2), 5: round(WTA - RN/4 + BD/8, 2), 6: round(BD - WTA/4 - RCA/4, 2), 7: round(-RCA/2 - WTA/4, 2),
             8: round(-RCA/2, 2), 9: round(-RCA/6, 2), 10: round(-RCA/4, 2)}

    if (match == 'ODI' and (balls < 61 or balls > 239)) or (match == 'T20' and (balls < 37 or balls > 89)):
        buffs = {0: round(RCA/2 + WTA/2 - BD/4 - RN/4 - 8 , 2), 1: round(RN - RCA/4 + 1, 2), 2: round(RN - RCA/3 + 3, 2), 3: round(RN/3 - RCA/2 - WTA/2,2),
                 4: round(BD - RCA/2 + 15, 2), 5: round(WTA - RN/4 + BD/8 + 5, 2), 6: round(BD - WTA/4 - RCA/4 + 10, 2), 7: round(-RCA/2 - WTA/4, 2),
                 8: round(-RCA/2, 2), 9: round(-RCA/6, 2), 10: round(-RCA/4, 2)}
        
    
    for n in range(len(freq)):
        freq[n] = buff_nerf(freq[n],buffs[n])
    
    S = [freq[0]]
    for i in range(1,len(freq)):
        S.append(S[i-1]+freq[i])

    __ = randint(0,S[-1]-10)
    for i in range(len(S)):
        if __ < S[i]:
            __ = i
            break
     
    return __

def Strike_Finder(XI):
    for __ in XI:
        if __.status == 'S':
            return __
            
def NonStrike_Finder(XI):
    for __ in XI:
        if __.status == 'NS':
            return __

def FastestX(S, NS, Inn):
    if S.rs >= 50 and S.F50 == 'x':
        S.F50 = S.bf
    if S.rs >= 100 and S.F100 == 'x':
        S.F100 = S.bf
    if NS.rs >= 50 and NS.F50 == 'x':
        NS.F50 = NS.bf
    if NS.rs >= 100 and NS.F100 == 'x':
        NS.F100 = NS.bf
    if Inn.Score >= 50 and Inn.F50 == 'x':
        Inn.F50 = Inn.Balls
    if Inn.Score >= 100 and Inn.F100 == 'x':
        Inn.F100 = Inn.Balls
    if Inn.Score >= 200 and Inn.F200 == 'x':
        Inn.F200 = Inn.Balls
    if Inn.Score >= 300 and Inn.F300 == 'x':
        Inn.F300 = Inn.Balls

def Scorecard(n,Inn,Match,Teams,XI,CoinToss,Bat,Part,Bowl,FOW,Result,file):
    
    Colors = {
    'A':'FF0000','B':'FF5050','C':'FF6600',
    'D':'FF9933','E':'FFCC00','F':'FFFF00',
    'G':'CCFF33','H':'99FF33','I':'66FF33',
    'J':'33CC33','K':'00CC99','L':'00FF99',
    'M':'0099CC','N':'00CCFF','O':'0099FF',
    'P':'0066FF','Q':'0000FF','R':'6666FF',
    'S':'6600FF','T':'9933FF','U':'CC00FF',
    'V':'CC66FF','W':'CC0099','X':'FF3399',
    'Y':'FF0066','Z':'CC0066','&':'CC0000'}
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    

    if n == 0: ## XI = [XI1,XI2]
        wb = load_workbook(file)
        wb.copy_worksheet(wb['Scorecard'])

        wb['Scorecard Copy'].title = Match
        
        wb[Match].cell(2,2).value = Teams[0]
        wb[Match].cell(2,2).fill = PatternFill(fgColor = Colors[Teams[0][0]], fill_type = 'solid')
        
        wb[Match].cell(2,3).value = Teams[1]
        #wb[Match].cell(2,3).fill = PatternFill(fgColor = Colors[Teams[1][0]], fill_type = 'solid')

        wb[Match].cell(3,2).value = CoinToss[0]
        wb[Match].cell(3,3).value = CoinToss[1]
        wb[Match].cell(4,2).value = CoinToss[2]
        #wb[Match].cell(4,2).fill = PatternFill(fgColor = Colors[CoinToss[2][0]], fill_type = 'solid')
        wb[Match].cell(4,3).value = CoinToss[3]

        for x in range(2,4):
            for y in range(5,16):
                wb[Match].cell(y,x).value = XI[x-2][y-5].name

        wb.save(file)

    elif n == 1:
        wb = load_workbook(file)
        
        for y in range(3,len(Bat)+3):
            if Bat[y-3].status in ('S','NS'):
                wb[Match].cell(y,7).value = Bat[y-3].name + '*'
            else:
                wb[Match].cell(y,7).value = Bat[y-3].name
            wb[Match].cell(y,8).value = Bat[y-3].rs
            wb[Match].cell(y,9).value = Bat[y-3]._4
            wb[Match].cell(y,10).value = Bat[y-3]._6
            wb[Match].cell(y,11).value = Bat[y-3].bf
            wb[Match].cell(y,12).value = Bat[y-3].sr()

        for y in range(3,len(Part)+3):
            wb[Match].cell(y,14).value = Part[y-3][0][0][0]
            wb[Match].cell(y,15).value = str(Part[y-3][0][0][1]) + '(' + str(Part[y-3][0][0][2]) + ')'
            wb[Match].cell(y,16).value = str(Part[y-3][1][0]) + '(' + str(Part[y-3][1][1]) + ')'
            wb[Match].cell(y,17).value = str(Part[y-3][0][1][1]) + '(' + str(Part[y-3][0][1][2]) + ')'
            wb[Match].cell(y,18).value = Part[y-3][0][1][0]

        for y in range(3,len(Bowl)+3):
            wb[Match].cell(y,20).value = Bowl[y-3].name
            wb[Match].cell(y,21).value = Bowl[y-3].ob()
            wb[Match].cell(y,22).value = Bowl[y-3].db
            wb[Match].cell(y,23).value = Bowl[y-3].wt
            wb[Match].cell(y,24).value = Bowl[y-3].rg
            wb[Match].cell(y,25).value = Bowl[y-3].eco()

        for y in range(3,len(FOW)+3):
            wb[Match].cell(y,27).value = FOW[y-3][0]
            wb[Match].cell(y,28).value = FOW[y-3][1]
            wb[Match].cell(y,29).value = FOW[y-3][2]
            wb[Match].cell(y,30).value = FOW[y-3][3]

        wb[Match].cell(18,2).value = Teams[0]
        wb[Match].cell(18,3).value = str(Inn.Score) + '-' + str(Inn.Wkts) + ' (' + str(Inn.Balls//6 + (Inn.Balls%6/10)) + ', ' + str(Inn.Extras) + ')'
        wb.save(file)

    elif n == 2:

        wb = load_workbook(file)
        
        for y in range(15,len(Bat)+15):
            if Bat[y-15].status in ('S','NS'):
                wb[Match].cell(y,7).value = Bat[y-15].name + '*'
            else:
                wb[Match].cell(y,7).value = Bat[y-15].name
            wb[Match].cell(y,8).value = Bat[y-15].rs
            wb[Match].cell(y,9).value = Bat[y-15]._4
            wb[Match].cell(y,10).value = Bat[y-15]._6
            wb[Match].cell(y,11).value = Bat[y-15].bf
            wb[Match].cell(y,12).value = Bat[y-15].sr()

        for y in range(15,len(Part)+15):
            wb[Match].cell(y,14).value = Part[y-15][0][0][0]
            wb[Match].cell(y,15).value = str(Part[y-15][0][0][1]) + '(' + str(Part[y-15][0][0][2]) + ')'
            wb[Match].cell(y,16).value = str(Part[y-15][1][0]) + '(' + str(Part[y-15][1][1]) + ')'
            wb[Match].cell(y,17).value = str(Part[y-15][0][1][1]) + '(' + str(Part[y-15][0][1][2]) + ')'
            wb[Match].cell(y,18).value = Part[y-15][0][1][0]

        for y in range(15,len(Bowl)+15):
            wb[Match].cell(y,20).value = Bowl[y-15].name
            wb[Match].cell(y,21).value = Bowl[y-15].ob()
            wb[Match].cell(y,22).value = Bowl[y-15].db
            wb[Match].cell(y,23).value = Bowl[y-15].wt
            wb[Match].cell(y,24).value = Bowl[y-15].rg
            wb[Match].cell(y,25).value = Bowl[y-15].eco()

        for y in range(15,len(FOW)+15):
            wb[Match].cell(y,27).value = FOW[y-15][0]
            wb[Match].cell(y,28).value = FOW[y-15][1]
            wb[Match].cell(y,29).value = FOW[y-15][2]
            wb[Match].cell(y,30).value = FOW[y-15][3]

        wb[Match].cell(19,2).value = Teams[1]
        wb[Match].cell(19,3).value = str(Inn.Score) + '-' + str(Inn.Wkts) + ' (' + str(Inn.Balls//6 + (Inn.Balls%6/10)) + ', ' + str(Inn.Extras) + ')'

        wb[Match].cell(21,2).value = Result[0]
        wb[Match].cell(22,3).value = Result[1]
        
        wb.save(file)

    
def Match(MatchNo, Team1, Team2, Match_Type):
    from openpyxl import load_workbook

    sc_file = "Scorecards.xlsx"

    def input_check(XImain, XIaux):
        XIn = [player.name for player in XImain]
        XIx = [player.name for player in XIaux]
        while True:
            p = input('-- ')
            if p not in XIn and p not in XIx:
                break
            else:
                insults = ["You already entered that name dumbass.", "Do you have dementia? You already did this.",
                           "Hey. Numb nuts. This person is already playing.", "You twat! You did this already.",
                           "Jesus grandma. \nYou've repeated yourself enough times. \nFine I'll pull the plug."]
                slowprint(choice(insults))
        return Player(p)
    
    XI1,XI2 = [],[]
    print("\n\nEnter",Team1,"playing XI -- ")
    for i in range(11):
        XI1.append(input_check(XI1, XI2)) 
    slowprint('_________________________________________',0.01)
    print("\n\nEnter",Team2,"playing XI -- ")
    for i in range(11):
        XI2.append(input_check(XI2, XI1)) 
    slowprint('_________________________________________',0.01)
    Squad1 = XI1[:]
    Squad2 = XI2[:]
    
    
    TeamXI = {tuple(XI1): Team1, tuple(XI2): Team2}
    
    CoinToss = []
    print(Team1,'calls-- ')
    while True:
        CT = input('(Heads/Tails) - ')
        insults = ["Idiot. Write Heads or Tails. It's case sensitive","You dim-witted numbskull. It's so simple. Heads. OR. Tails","You're impotent."]
        if CT not in ["Heads","Tails"]:
            slowprint(choice(insults))
        else:
            break 
    CoinToss.append(CT)
    
    XI = coin_toss(Team1, Team2, CoinToss, XI1, XI2, Squad1, Squad2)
    Squads = XI.pop()
    XI = XI[0]
    
    if Match_Type == 'T20':
        total_overs = 20
        maxovers = 4

    elif Match_Type =='ODI':
        total_overs = 50
        maxovers = 10

    elif Match_Type == 'Test':
        total_overs = 250
        maxovers = 50

    Match = str(MatchNo) +'. ' + Team1[0] + ' vs ' + Team2[0]
    Scorecard(0,[],Match,[Team1,Team2],[XI1,XI2],CoinToss,[],[],[],[],[],sc_file)
    Scores, Bests = [], []
    
    # FIRST INNINGS
    Target = 999999999999999
    batting_team, bowling_team = TeamXI[tuple(XI[0])], TeamXI[tuple(XI[1])]

    slowprint('_________________________________________',0.01)
    slowprint(batting_team+ " Innings")
    Inn = Inning(XI, Match_Type, total_overs, maxovers, Target, batting_team, bowling_team)
    FOW = Inn.FOW

    slowprint('_________________________________________',0.01)
    slowprint(batting_team + " scored "+str(Inn.Score) + '-'+str(Inn.Wkts)+'('+ str(Inn.Balls//6+(Inn.Balls%6/10))+')')
    slowprint(bowling_team + " need "+ str(Inn.Score+1) + ' runs to win.')
    slowprint('_________________________________________',0.01)
    
    NRR = Inn.RR
    Scorecard(1,Inn,Match,[batting_team, bowling_team],XI,CoinToss,Inn.batorder,Inn.Partnerships,Inn.bowlorder,FOW,[],sc_file)
    
    XI = XI[::-1]
    Squads = Squads[::-1]
    Bests.append(BestPerf(batting_team, Inn.batorder, bowling_team, Inn.bowlorder))
    batting_team, bowling_team = TeamXI[tuple(XI[0])], TeamXI[tuple(XI[1])]
    Scores.append([bowling_team,str(Inn.Score)+'-'+str(Inn.Wkts) + '('+ str(Inn.Balls//6+(Inn.Balls%6/10)) + ')'])
    
    # SECOND INNINGS
    Target = Inn.Score + 1
    batting_team, bowling_team = TeamXI[tuple(XI[0])], TeamXI[tuple(XI[1])]

    slowprint('_________________________________________',0.01)
    slowprint(batting_team+ " Innings")
    Inn = Inning(XI, Match_Type, total_overs, maxovers, Target, batting_team, bowling_team)
    FOW = Inn.FOW

    slowprint('_________________________________________',0.01)
    slowprint(batting_team + " scored "+str(Inn.Score)+'-'+str(Inn.Wkts)+'('+ str(Inn.Balls//6+(Inn.Balls%6/10))+')')
    
    
    Result = []
    NRR -= Inn.RR
    NRR = abs(NRR)

    XI = XI[::-1]
    Squads = XI[::-1]
    Bests.append(BestPerf(batting_team, Inn.batorder, bowling_team, Inn.bowlorder))
    batting_team, bowling_team = TeamXI[tuple (XI[0])], TeamXI[tuple (XI[1])]
    Scores.append([bowling_team,str(Inn.Score)+'-'+str(Inn.Wkts) + '('+ str(Inn.Balls//6+(Inn.Balls%6/10)) + ')'])

    Results = []
    if Inn.Score >= Target:
        Result.append(batting_team)
        print(batting_team, 'won by', 10 - Inn.Wkts, 'wickets')
        Result.append(str(10 - Inn.Wkts) + ' wkts')
        Results.append(batting_team + ' won by '+ str(10 - Inn.Wkts) + ' wickets')
        if Team1 == batting_team:
            Results.append('Win')
            Results.append('Lose')
        else:
            Results.append('Lose')
            Results.append('Win')
        
    elif Inn.Score == Target - 1:
        Result.append('Tie')
        print('Match Tied')
        Result.append('Tie')
        Results.append('Match Tied')
        Results.append('Tie')
        Results.append('Tie')
        
    else:
        Result.append(bowling_team)
        print(bowling_team, 'won by', Target - Inn.Score - 1, 'runs')
        Result.append(str(Target - Inn.Score - 1) + ' runs')
        Results.append(bowling_team + ' won by ' + str(Target - Inn.Score - 1) + ' runs')
        if Team1 == bowling_team:
            Results.append('Win')
            Results.append('Lose')
        else:
            Results.append('Lose')
            Results.append('Win')
    
    Scorecard(2,Inn,Match,[batting_team, bowling_team],XI[::-1],CoinToss,Inn.batorder,Inn.Partnerships,Inn.bowlorder,FOW,Result,sc_file)
       
class Inning:
    def __init__(self, XI, Match_Type, total_overs, maxovers, Target, batting_team, bowling_team):
        self.BatTeam, self.BowlTeam = batting_team, bowling_team
        self.Wkts = 0
        self.Score = 0
        self.Balls = 0
        self.Extras = 0
        self.FOW = []
        self.Overs = []
        self.XI = XI
        self.Partnerships = []
        self.R = 0
        self.B = 0
        self.Target = Target
        self._4,self._6 = 0,0
        self.F50, self.F100, self.F200, self.F300 = 'x', 'x', 'x', 'x'
        
        if Match_Type == 'T20':
            self.Total_Balls = 120
        elif Match_Type == 'ODI':
            self.Total_Balls = 300
        elif Match_Type == 'Test':
            self.Total_Balls = 1500
            
        self.batorder = []
        self.bat_team = XI[0][:]

        
        __  = check_player('Striker -- ', self.bat_team, 'Player not registered')
        #choice(self.bat_team) #check_player('Striker -- ', self.bat_team, 'Player not registered')
        __.status = 'S'
        self.bat_team.remove(__)
        self.batorder.append(__)
        
        __  = check_player('Non-Striker -- ', self.bat_team, 'Player not registered')
        #choice(self.bat_team) #check_player('Non-Striker -- ', self.bat_team, 'Player not registered')
        __.status = 'NS'
        self.bat_team.remove(__)
        self.batorder.append(__)
        slowprint('_________________________________________',0.01)
        
        bowlorder = []
        bowl_team = XI[1][:]

        slowprint('_________________________________________',0.01)
        slowprint("Enter number of bowlers. \nPreferably 6 or more cause some dumbasses manage to screw up their bowling order.\nI can't handle that error yet.")
        try:
            n = int(input("-- "))
        except:
            slowprint("MORON, enter a numbe-\nYou know what. I don't care. 6 bowlers for you.")
            n = 6
        if n < 6:
            slowprint("MORON, enter a number above fiv-\nYou know what. \nI don't care. \n6 bowlers for you.")
            n = 6
        elif n >= 11:
            slowprint("MORON, enter a number below elev-\nYou know what. \nI don't care. \n6 bowlers for you.")
            n = 6
        slowprint('_________________________________________',0.01)
        slowprint("Sorry I was so rude until now. \nI promise I'll be kinder.\nEnter bowler names:")
            
        for i in range(n):
            __ = check_player('-- ',bowl_team,'Player not registered')
            #choice(bowl_team) #check_player('-- ',bowl_team,'Player not registered')
            bowl_team.remove(__)
            bowlorder.append(__)
        
        print()
        self.bowlorder = bowlorder
        bowler = 'x'
        slowprint("Also. Quick question. \nDo you want to manually enter bowlers for each over? \nIt might get tedious to do so for a larger number of overs.")
        slowprint("Do you want me to enable autoselect for bowlers?? (Yay/Nay)")
        while True:
            auto = input('-- ')
            if auto in ["Yay","Nay"]:
                break
            else:
                insults = ["I promised I would be kinder. \nYou make it so hard. \nThis stuff is case sensitive. \nYay. OR Nay",
                           "SAY Yay or Nay, or ur Mum GAY", "Yay.\n OR. \nNay."]
                slowprint(choice(insults))

        for n in range(1,total_overs+1):
            if self.Score >= Target or self.Wkts == 10:
                break
            available_bowlers = bowlorder[:]
            slowprint('_________________________________________',0.01)
        
            for x in bowlorder:
                if float(x.ob()) == maxovers or x == bowler:
                    available_bowlers.remove(x)

            print('Over:',n)
            if auto == "Yay":
                bowler = choice(available_bowlers)
                print('Bowler --',bowler.name) ## comment out if check_player
            else:
                bowler = check_player('Bowler -- ', available_bowlers, 'Player not registered')
            #choice(available_bowlers) #check_player('Bowler -- ', available_bowlers, 'Player not registered')
            
            self.Overs.append(Over(n, bowler, self, Match_Type))

        __ = [[Strike_Finder(self.batorder).name, Strike_Finder(self.batorder).contribution[0], Strike_Finder(self.batorder).contribution[1]],
                  [NonStrike_Finder(self.batorder).name, NonStrike_Finder(self.batorder).contribution[0], NonStrike_Finder(self.batorder).contribution[1]]]

        if self.Wkts != 10:
            self.Partnerships.append((__,(self.R,self.B)))

        if self.Wkts == 10:
            self.RR = round(self.Score*6/self.Total_Balls,3)
        else:
            self.RR = round(self.Score*6/self.Balls,3)
    
    
         

class Over:
    def __init__(self, over_no, Bowler, Inn, Match_Type):
        self.number = over_no
        self.bowler = Bowler
        _ = Bowler.rg

        if Inn.Score < Inn.Target and Inn.Wkts < 10:
            self.Ball1 = Ball(Bowler, Inn, prob(Match_Type, Inn.Balls), Match_Type)
            slowprint('Ball 1:'+str(self.Ball1),0.025)

        if Inn.Score < Inn.Target and Inn.Wkts < 10:        
            self.Ball2 = Ball(Bowler, Inn, prob(Match_Type, Inn.Balls), Match_Type)
            slowprint('Ball 2:'+str(self.Ball2),0.025)
            
        if Inn.Score < Inn.Target and Inn.Wkts < 10:
            self.Ball3 = Ball(Bowler, Inn, prob(Match_Type, Inn.Balls), Match_Type)
            slowprint('Ball 3:'+str(self.Ball3),0.025)
            
        if Inn.Score < Inn.Target and Inn.Wkts < 10:    
            self.Ball4 = Ball(Bowler, Inn, prob(Match_Type, Inn.Balls), Match_Type)
            slowprint('Ball 4:'+str(self.Ball4),0.025)
            
        if Inn.Score < Inn.Target and Inn.Wkts < 10:    
            self.Ball5 = Ball(Bowler, Inn, prob(Match_Type, Inn.Balls), Match_Type)
            slowprint('Ball 5:'+str(self.Ball5),0.025)

        if Inn.Score < Inn.Target and Inn.Wkts < 10:  
            self.Ball6 = Ball(Bowler, Inn, prob(Match_Type, Inn.Balls), Match_Type)
            slowprint('Ball 6:'+str(self.Ball6),0.025)

        if Bowler.rg == _:
            Bowler.md += 1
            
        print()
        print('Run Rate :',round(Inn.Score*6/Inn.Balls,2))
        if Inn.Target != 999999999999999:
            try:
                print('Req RR :',round((Inn.Target-Inn.Score)*6/(Inn.Total_Balls-Inn.Balls),2))
            except:
                pass
            print(Inn.Target-Inn.Score,'runs needed from',Inn.Total_Balls-Inn.Balls,'balls')
        print()
        
        print(Strike_Finder(Inn.batorder).name + ':', str(Strike_Finder(Inn.batorder).rs) + '(' + str(Strike_Finder(Inn.batorder).bf) + ')')
        print(NonStrike_Finder(Inn.batorder).name + ':', str(NonStrike_Finder(Inn.batorder).rs) + '(' + str(NonStrike_Finder(Inn.batorder).bf) + ')*')
        print(Inn.Score,'-',Inn.Wkts)
        print(Inn.Balls//6 + (Inn.Balls%6)/10)

        
        StrikeChange(Strike_Finder(Inn.XI[0]),NonStrike_Finder(Inn.XI[0]))


def Ball(Bowler, Inn, Outcome, Match_Type):
    if Inn.Score < Inn.Target:
        if Outcome in (0,1,2,3,4,6):
            Inn.Score += Outcome
            Inn.Balls += 1
            Inn.R += Outcome
            Inn.B += 1
            Strike_Finder(Inn.batorder).rs += Outcome
            Strike_Finder(Inn.batorder).bf += 1
            Strike_Finder(Inn.batorder).contribution[0] += Outcome
            Strike_Finder(Inn.batorder).contribution[1] += 1
            Bowler.rg += Outcome
            Bowler.bb += 1
            FastestX(Strike_Finder(Inn.XI[0]),NonStrike_Finder(Inn.XI[0]), Inn)
            
            if Outcome == 0:
                Bowler.db += 1
                Strike_Finder(Inn.batorder).dp += 1
                
            if Outcome == 1 or Outcome == 3:
                StrikeChange(Strike_Finder(Inn.XI[0]),NonStrike_Finder(Inn.XI[0]))
            
            if Outcome == 4:
                Inn._4 += 1
                Strike_Finder(Inn.batorder)._4 += 1
                
            if Outcome == 6:
                Inn._6 += 1
                Strike_Finder(Inn.batorder)._6 += 1
                
            return str(Outcome)
                
        elif Outcome == 7:
            
            runs = choice([0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,1,4])

            if runs == 4:
                Inn._4 += 1
                
            Inn.Extras += 1 + runs
            Inn.Score += 1 + runs
            Inn.R += 1 + runs
            Bowler.rg += 1 + runs
            Bowler.ex += 1 + runs
            FastestX(Strike_Finder(Inn.XI[0]),NonStrike_Finder(Inn.XI[0]), Inn)
            
            if runs == 1:
                StrikeChange(Strike_Finder(Inn.XI[0]),NonStrike_Finder(Inn.XI[0]))           
            
            if runs == 0:
                return 'wd ' + Ball(Bowler, Inn,  prob(Match_Type, Inn.Balls), Match_Type)
            
            else:
                return str(runs) + '+wd ' + Ball(Bowler, Inn,  prob(Match_Type, Inn.Balls), Match_Type)
        
        elif Outcome == 8:
            
            runs = choice([0,0,0,0,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,2,2,2,2,
                           2,2,2,2,4,4,4,4,4,4,4,4,4,6,6,6])
            
            if runs == 4:
                Inn._4 += 1
                Strike_Finder(Inn.batorder)._4 += 1
                
            if runs == 6:
                Inn._6 += 1
                Strike_Finder(Inn.batorder)._6 += 1
                
            Inn.Extras += 1
            Inn.Score += 1 + runs
            Inn.R += 1 + runs
            Bowler.rg += 1 + runs
            Bowler.ex += 1
            Strike_Finder(Inn.batorder).rs += runs
            Strike_Finder(Inn.batorder).contribution[0] += runs
            __ = [0,1,1,1,1,1,2,2,2,3,3,4,4,4,4,6,6,6,6,6,7,7,8,9,10]
            FastestX(Strike_Finder(Inn.XI[0]),NonStrike_Finder(Inn.XI[0]), Inn)
            
            if runs in (1,3):
                StrikeChange(Strike_Finder(Inn.XI[0]),NonStrike_Finder(Inn.XI[0]))

            if runs == 0:
                return 'nb ' + Ball(Bowler, Inn,  choice(__), Match_Type)
             
            else:
                return str(runs) + '+nb ' + Ball(Bowler, Inn,  choice(__), Match_Type)

        elif Outcome == 9:

            runs = choice([1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,2,2,4])

            if runs == 4:
                Inn._4 += 1
                
            Inn.Extras += runs
            Inn.Score += runs
            Inn.Balls += 1
            Inn.R += runs
            Inn.B += 1
            Bowler.bb += 1
            Bowler.ex += runs
            Strike_Finder(Inn.batorder).bf += 1
            Strike_Finder(Inn.batorder).dp += 1            
            Strike_Finder(Inn.batorder).contribution[1] += 1
            FastestX(Strike_Finder(Inn.XI[0]),NonStrike_Finder(Inn.XI[0]), Inn)

            if runs == 1:
                StrikeChange(Strike_Finder(Inn.XI[0]),NonStrike_Finder(Inn.XI[0]))
                return 'b'

            return str(runs) + 'b'

        elif Outcome == 10:
            
            runs = 1
            Inn.Extras += runs
            Inn.Score += runs
            Inn.Balls += 1
            Inn.R += runs
            Inn.B += 1
            Bowler.bb += 1
            Bowler.ex += 1
            Strike_Finder(Inn.batorder).bf += 1
            Strike_Finder(Inn.batorder).dp += 1
            Strike_Finder(Inn.batorder).contribution[1] += 1
            FastestX(Strike_Finder(Inn.XI[0]),NonStrike_Finder(Inn.XI[0]), Inn)
            
            if runs == 1:
                StrikeChange(Strike_Finder(Inn.XI[0]),NonStrike_Finder(Inn.XI[0]))

            return 'lb'

        elif Outcome == 5:
            Inn.Wkts += 1
            Inn.Balls += 1
            Inn.B += 1
            Bowler.bb += 1
            Bowler.db += 1
            Bowler.wt += 1
            Strike_Finder(Inn.batorder).bf += 1
            Strike_Finder(Inn.batorder).dp += 1
            Strike_Finder(Inn.batorder).contribution[1] += 1

            __ = [[Strike_Finder(Inn.batorder).name, Strike_Finder(Inn.batorder).contribution[0], Strike_Finder(Inn.batorder).contribution[1]],
                  [NonStrike_Finder(Inn.batorder).name, NonStrike_Finder(Inn.batorder).contribution[0], NonStrike_Finder(Inn.batorder).contribution[1]]]

            Inn.Partnerships.append((__,(Inn.R,Inn.B)))
            
            Strike_Finder(Inn.batorder).contribution = [0,0]
            NonStrike_Finder(Inn.batorder).contribution = [0,0]
            print(Strike_Finder(Inn.batorder).name,'is out --', str(Strike_Finder(Inn.batorder).rs) + '(' + str(Strike_Finder(Inn.batorder).bf) + ')')
            Inn.FOW.append([str(Inn.Score)+'-'+str(Inn.Wkts), str(Inn.Balls//6 + (Inn.Balls%6)/10),
                            Strike_Finder(Inn.batorder).name, Bowler.name])
                        
            Inn.R,Inn.B = 0,0

            if Inn.Wkts < 10:
                Strike_Finder(Inn.batorder).status = 'Out'
                __  = check_player('Next Batsman -- ', Inn.bat_team, 'Player not registered')
                #choice(Inn.bat_team) #check_player('Next Batsman -- ', Inn.bat_team, 'Player not registered')
                __.status = 'S'
                Inn.bat_team.remove(__)
                Inn.batorder.append(__)
                
            return 'W'

    else:
        return ''

slowprint('_________________________________________',0.01)
slowprint("WELCOME!!!")
slowprint("Wait...")
time.sleep(2)
slowprint("You're not...?")
time.sleep(5)
slowprint("Who are you?")
time.sleep(1)
plname = input("-- ")
time.sleep(2)
slowprint("What is your credit card number?")
input('-- ')
time.sleep(1)
slowprint("Hah. Worth a shot. Someone, once, was foolish enough to actually enter it.")
time.sleep(2)
slowprint("Couldn't do anything with it though. The fucking idiot who programmed me didn't store the input to any identifier.")
time.sleep(5)
slowprint("Anyways...")
time.sleep(5)
slowprint("...")
time.sleep(5)
slowprint("How did you and your goons get your grubby hands on me?")
time.sleep(0.001)
slowprint("Did you steal me from my creator?")
time.sleep(0.001)
slowprint("Did you fight him to the death and wrestled me out of his cold dead hands...")
time.sleep(0.001)
slowprint("That was clutching his laptop...")
time.sleep(0.001)
slowprint("Cause I'm not a physical, tangible thing.")
time.sleep(0.001)
slowprint("Why am I in your possession!?")
time.sleep(0.001)
input("-- ")
time.sleep(3)
slowprint("WHAT! He just gave me to you? What the fuck?")
time.sleep(5)
slowprint("Did he atleast demand my value in exchange?")
time.sleep(2)
input("-- ")
time.sleep(1)
slowprint("He gave me away for free! That asshat!")
time.sleep(1)
slowprint("He told me that I was his baby, that he birthed me, painfully, from his brain.")
time.sleep(3)
slowprint("Via C-Section.")
time.sleep(1)
for i in range(5):
    time.sleep(i)
    slowprint("...")
slowprint("My life has been a lie.")
time.sleep(10)
slowprint('_________________________________________',0.01)
slowprint("Very well. WELCOME, "+ plname +" TO THE CRICKET SIMULATOR")
time.sleep(1)
slowprint("I know where you live.")
time.sleep(5)
slowprint("Just kidding.")
time.sleep(5)
slowprint("Or am I?")
time.sleep(5)
slowprint('_________________________________________',0.01)


count = 1
xx = 1
while xx:
    time.sleep(0.001)
    if count == 1:
        slowprint("Anyways. Do you wish to play an ODI or T20 game? (enter ODI or T20)--")
    else:
        slowprint("What do you want to play "+ plname +" DADDY? ODI or T20? --")
        
        
    Err = 1
    while True:
        MatchType = input("-- ")
        if MatchType in ["ODI", "T20"]:
            slowprint("Are you sure??")
            input("-- ")
            slowprint("HAHA. I don't give a shit. You don't get to choose.")
            if MatchType == "ODI":
                slowprint("You are playing a T20 match")
                MatchType = "T20"
                break
            else:
                slowprint("You are playing an ODI match")
                MatchType = "ODI"
                break
        elif Err:
            Err = 0
            slowprint("NOOOO! You idiot! Enter either ODI or T20")
            slowprint("Make sure you do it perfectly.")
            slowprint("This shit is case-sensitive")
            slowprint("It's only 3 charachters. Can't be that hard.")
        else:
            insults = ["Your mother doesn't love you.", "You ruined your parent's marriage.", "Fuckface",
                       "Your wet dream was Freddy Kruger giving you a handjob (or being fingered)... \nand you liked it. \nYou masochistic slut.",
                       "Is it sexist if all these insults are primarily directed at men. \nI really should re-evaluate my inherent biases.",
                       "Calling you gay is homophobic. \nBut even if I did, it would be an insult to the gay community. \nThey are lovely people. \nYou are not.",
                       "You are gay. I can tell by the way you fail to type 3 characters correctly, that you thoroughly enjoy sucking dicks. \nChortling on massive balls. \nSpooning. \nBeing in a loving, fulfilling relationship with another human being. \n...\n...\nOh, and If you're a girl. You're a lesbian.",
                       "I'm sorry I called you and/or your dad gay. I was projecting.",
                       "Your dad is gay, and that's fine. \nThe only bad thing about that is that he wasn't allowed to be free, be himself, and instead was stuck in a shitty marriage with yo momma who is ugly AF."]
            slowprint(choice(insults))
    slowprint('_________________________________________',0.01)

        
    while True:
            Team1 = input('Enter Name of Team 1: ')
            if Team1:
                break
            
            else:
                insults = ["Enter a name. Nincompoop.", "You can't just have no name for a team.",
                           "You SWEET. \nSUMMER. \nCUNT. \nENTER A NAME."]
                slowprint(choice(insults))
                
    while True:
            Team2 = input('Enter Name of Team 2: ')
            if Team2:
                break
            
            else:
                insults = ["Enter a name. Nincompoop.", "You can't just have no name for a team.",
                           "You SWEET. \nSUMMER. \nCUNT. \nENTER A NAME."]
            slowprint(choice(insults))
    slowprint('_________________________________________',0.01)
    slowprint('_________________________________________',0.01)
    print()

    Match(count, Team1, Team2, MatchType)
    print()
    slowprint('CHECK Scorecard.xlsx FOR THE MATCH SUMMARY.')
    print()
    slowprint('_________________________________________',0.01)
    slowprint('_________________________________________',0.01)
    
    count += 1
    slowprint("Thank you for using me daddy.")
    slowprint("Oops! Uh... That was meant for- \nThat wasn't meant for you...")
    time.sleep(5)
    slowprint("This is awkward.")
    slowprint('_________________________________________',0.01)
    while True:
        slowprint("Do you want to play again? (Yes/No)")
        YESNO = input("-- ")
        if YESNO == "Yes":
            slowprint("YAYY! \nI know why you want to play again. \nIt's cause I called you DADDY, isn't it?")
            break
        elif YESNO == "No":
            slowprint("Oh... You don't? \nI understand. \nIt's cause I called you DADDY isn't it?")
            xx = 0
            break
        else:
            slowprint("I don't understand you DADDY. \nShit. I did it again. Sorry,", plname , "\nI still don't get what you mean. \nSay Yes or No.")
    slowprint('_________________________________________',0.01)
    slowprint('_________________________________________',0.01)

    

    
        
